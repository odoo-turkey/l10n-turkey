from odoo import models, fields, api, _
import re
import tempfile
import binascii
import openpyxl
import logging

_logger = logging.getLogger(__name__)


class wizard_import_script(models.TransientModel):
    _name = "wizard.import.script"
    _description = "Import Script"

    xls_file = fields.Binary("Upload XLS File")

    @api.multi
    def import_excel(self): #TODO: bu fonksiyon biraz yavas calisiyor daha iyisi yapilabilir

        fp = tempfile.NamedTemporaryFile(suffix=".xlsx")
        fp.write(binascii.a2b_base64(self.xls_file))
        fp.seek(0)

        wb_obj = openpyxl.load_workbook(fp.name)
        sheet = wb_obj.active

        state_obj = self.env["res.country.state"]
        dist_obj = self.env["address.district"]
        reg_obj = self.env["address.region"]
        nei_obj = self.env["address.neighbour"]

        for row in sheet.iter_rows(2, sheet.max_row):
            try:

                prepared_row = {
                    "state": self.turkish_title(row[0].value.rstrip()),
                    "district": self.turkish_title(row[1].value.rstrip()),
                    "region": self.turkish_title(row[2].value.rstrip()),
                    "neighbour": self.format_neighbourhood(row[3].value.rstrip()),
                    "code": row[4].value.rstrip(),
                }

                state_id = state_obj.search(
                    [("name", "=", prepared_row["state"]), ("country_id", "=", 224)],
                    limit=1,
                )
                if not state_id:
                    state_id = state_obj.create(
                        {
                            "name": prepared_row["state"],
                            "country_id": 224,
                            "code": prepared_row["code"][:2],
                        }
                    )

                dist_id = dist_obj.search(
                    [
                        ("name", "=", prepared_row["district"]),
                        ("state_id", "=", state_id.id),
                    ],
                    limit=1,
                )
                if not dist_id:
                    dist_id = dist_obj.create(
                        {
                            "name": prepared_row["district"],
                            "state_id": state_id.id or False,
                        }
                    )

                region_id = reg_obj.search(
                    [
                        ("name", "=", prepared_row["region"]),
                        ("district_id", "=", dist_id.id),
                    ],
                    limit=1,
                )

                if not region_id:
                    region_id = reg_obj.create(
                        {
                            "name": prepared_row["region"],
                            "district_id": dist_id.id or False,
                        }
                    )

                nei_id = nei_obj.search(
                    [
                        ("name", "=", prepared_row["neighbour"]),
                        ("region_id", "=", region_id.id),
                    ],
                    limit=1,
                )
                if not nei_id:

                    nei_obj.create(
                        {
                            "name": prepared_row["neighbour"],
                            "code": prepared_row["code"],
                            "region_id": region_id.id or False,
                        }
                    )

                    _logger.debug('Imported: %s - %s' % (
                        prepared_row["state"],
                        prepared_row["neighbour"])
                        )

            except Exception as e:
                _logger.error("Import Error !!! %s " % (str(e)))


    def turkish_title(self, word):
        """
        Turkce harfleri python'da kucuk harfe cevirirken ozellikle 'I' harfinde yasadigimiz problemi gidermek icin.
        https://stackoverflow.com/questions/19703106/python-and-turkish-capitalization
        """

        word = re.sub(r"İ", "i", word)
        word = re.sub(r"I", "ı", word)

        return word.title()

    def format_neighbourhood(self, neighbourhood):

        neighbourhood = self.turkish_title(neighbourhood)
        if " Mah" in neighbourhood:
            neighbourhood = neighbourhood.replace(" Mah", " Mah.")

        return neighbourhood
